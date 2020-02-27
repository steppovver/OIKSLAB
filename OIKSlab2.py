# Rubens_COOL
from getch import pause
from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active


def lineLetter(line):
    a = [('А', 0), ('Б', 0), ('В', 0), ('Г', 0), ('Д', 0), ('Е', 0), ('Ё', 0), ('Ж', 0), ('З', 0), ('И', 0), ('Й', 0), ('К', 0), ('Л', 0), ('М', 0), ('Н', 0), ('О', 0), ('П', 0), ('Р', 0), ('С', 0), ('Т', 0), ('У', 0), ('Ф', 0), ('Х', 0), ('Ц', 0), ('Ч', 0), ('Ш', 0), ('Щ', 0), ('Ъ', 0), ('Ы', 0), ('Ь', 0), ('Э', 0), ('Ю', 0), ('Я', 0), ('.', 0), (',', 0), (' ', 0)]
    for letter in line:
        if letter != '\n':
            print(letter, end="")
            for i in range(len(a)):
                if letter == a[i][0] or letter == a[i][0].swapcase():
                    a[i] = (letter.upper(), a[i][1] + 1)
                    break
    return a


choise = int(input('Нажмите 1 чтобы ввести текст в консоль \nНажмите 2 чтобы достать текст из файла text.txt \n'))
if choise == 1:
    print("Enter/Paste your content. Ctrl-D or Ctrl-Z ( windows ) to save it.")
    fin = ''
    while True:
        try:
            line = input()
        except EOFError:
            break
        fin = fin + line
else:
    fin = open('text.txt', 'r', encoding='Windows-1251').read()
lengthF = len(fin)
cntword = len(fin.split())
print(cntword)
if cntword < 23:
    print('Недостаточное количество слов. Необходимо 23 слова как минимум')
    pause()
else:
    lengthF = lengthF // 24
    numLetter = 0
    for numLine in range(1, 24):
        cntLater = [0] * 36
        if numLine == 23:
            line = fin[numLetter:]
        else:
            line = fin[numLetter:lengthF+numLetter]
            numLetter = lengthF+numLetter
            while ord(line[-1]) != 32:
                line = line + fin[numLetter]
                numLetter += 1
        print(numLine, end=". ")
        cntLater = lineLetter(line)
        print()
        print(cntLater)
        ws.cell(row=numLine+1, column=1, value=numLine).font = Font(bold=True, color='DC143C')
        for i in range(len(cntLater)):
            ws.cell(row=numLine+1, column=i + 2, value=cntLater[i][1])
    for i in range(36):
        ws.cell(row=1, column=i+2, value=cntLater[i][0]).font = Font(bold=True, color='DC143C')
    wb.save('result.xlsx')
    pause()
